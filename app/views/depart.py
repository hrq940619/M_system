from django.http import HttpResponse
from django.shortcuts import render, redirect
from app import models
from app.utils.pagination import Pagination
from app.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm


# =============================== 部门管理 =====================================

def depart_list(request):
    """ 部门列表 """
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=5)
    context = {
        'queryset': page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, 'app/depart_list.html', context)


def depart_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, 'app/depart_add.html')
    # 获取用户POST提交过来的数据
    title = request.POST.get("title")
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向回部门列表
    return redirect("/depart_list/")


def depart_deleted(request):
    # 获取id
    depart_id = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=depart_id).delete()
    # 重定向回部门列表
    return redirect("/depart_list/")


def depart_edit(request, nid):
    """ 修改部门 """
    # 根据nid，获取数据
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id,row_object.title)

        return render(request, 'app/depart_edit.html', {"row_object": row_object})

    # 获取用户提交的标题
    title = request.POST.get("title")
    # 根据id找到数据库中的数据并更新
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向回部门列表
    return redirect("/depart_list/")


def depart_multi(request):
    """ 批量上传(Excel)文件 """
    from openpyxl import load_workbook

    # 1、获取用户上传的文件对象
    file_pbject = request.FILES.get('exc')
    # print(type(file_pbject))

    # 直接打开Excel并读取内容
    # 2、对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_pbject)
    sheet = wb.worksheets[0]

    # 3、循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart_list/')
